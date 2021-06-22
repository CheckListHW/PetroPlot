import math
import openpyxl

from tkinter import *

import numpy
import pandas as pd

import tkinter as tk
import tkinter.filedialog as fd

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

class SheetReader:
    Url = None
    Columns = []
    Old_workbook = None
    Sheet = None
    FZI = openpyxl.Workbook()
    FZI_sheet = None
    auto_FZI = None

    def __init__(self, url, depth_column_name, porv_column_name, pron_column_name, layer_column_name, note_column_name):
        self.depth_column_name = depth_column_name
        self.pron_column_name = pron_column_name
        self.porv_column_name = porv_column_name
        self.layer_column_name = layer_column_name
        self.note_column_name = note_column_name
        self.Url = url
        self.Old_workbook = openpyxl.load_workbook(url)
        self.Sheet = self.Old_workbook['1']

    def run(self):
        self.pull_columns_to_new_workbook([self.depth_column_name, self.porv_column_name, self.pron_column_name])
        self.calculate()

    def pull_columns_to_new_workbook(self, columns):
        row_number_old_sheet = 2
        row_number_new_sheet = 2
        self.FZI_sheet = self.FZI.active
        # очень нужно не удаляй
        columns_name = ['Глубина', 'Пористость', 'Проницаемость']
        for i in range(len(columns_name)):
            cell = self.FZI_sheet.cell(row=1, column=i + 1)
            cell.value = columns_name[i]

        while (self.Sheet[columns[0] + str(row_number_old_sheet)]).value:

            if self._row_valid(row_number_old_sheet):
                for i in range(len(columns)):
                    cell = self.FZI_sheet.cell(row=row_number_new_sheet, column=i + 1)
                    cell.value = self.Sheet[columns[i] + str(row_number_old_sheet)].value
                row_number_new_sheet += 1
            row_number_old_sheet += 1

        self.FZI.save('FZI.xlsx')

    def _row_valid(self, row_number: int) -> bool:
        if self.Sheet[self.depth_column_name + str(row_number)].value == \
                self.Sheet[self.depth_column_name + str(row_number - 1)].value:
            return False

        if str(self.Sheet[self.note_column_name + str(row_number)].value).__contains__('трещина'):
            return False

        if str(self.Sheet[self.porv_column_name + str(row_number)].value) == '0':
            return False
        if str(self.Sheet[self.porv_column_name + str(row_number)].value) == 'None':
            return False

        if str(self.Sheet[self.pron_column_name + str(row_number)].value) == '0':
            return False
        if str(self.Sheet[self.pron_column_name + str(row_number)].value) == 'None':
            return False

        for exception in ['ртинс', 'c1ok', 'c1bb', 'c1tl']:
            if self.Sheet[self.layer_column_name + str(row_number)].value.__contains__(exception):
                return False

        return True

    def calculate(self):
        df = pd.read_excel("FZI.xlsx", engine='openpyxl')
        df['Пористость'] = df['Пористость'] / 100
        df['porosity*'] = df['Пористость'] / (1 - df['Пористость'])
        df['корень'] = (df['Проницаемость'] / df['Пористость']) ** 0.5
        df['RQI'] = 0.0314 * df['корень']
        df['FZI'] = df['RQI'] / df['porosity*']
        df['Log(FZI)'] = numpy.log(df['FZI'])
        df['probability'] = 1 / (len(df))
        df = df.sort_values(by='Log(FZI)', ascending=True)  # сортировка данных по возрастанию
        df.to_excel('auto_FZI.xlsx')  # функция сохранения в эксель

        self.auto_FZI = openpyxl.load_workbook('auto_FZI.xlsx')
        auto_FZI_sheet = self.auto_FZI.active
        row_number = 3
        while auto_FZI_sheet['B' + str(row_number)].value:
            auto_FZI_sheet['J' + str(row_number)].value = \
                float(auto_FZI_sheet['J' + str(row_number - 1)].value) + \
                float(auto_FZI_sheet['J2'].value)
            row_number += 1
        self.auto_FZI.save('auto_FZI.xlsx')

    @staticmethod
    def get_column(filename, column_name) -> []:
        df = openpyxl.load_workbook(filename).active
        column = []
        for i in range(len(df[column_name])):
            column.append(df[column_name + str(i+1)].value)
        return column

    def get_column_auto_FZI(self, column_name) -> []:
        df = pd.read_excel("auto_FZI.xLsx", engine='openpyxl')
        return df[column_name]

    def get_columns_auto_FZI(self, columns_name) -> []:
        df = pd.read_excel("auto_FZI.xLsx", engine='openpyxl')
        return df[columns_name]

class App(tk.Tk):
    FZI_work_book_columns = ['AJ', 'X', 'AC', 'AP', 'AQ']

    FZI_chart_x = []
    FZI_chart_y = []

    rock_type_borders = []
    dots_rock_type = []
    pron_por_fzi = []

    rock_type_chart_scale = 'log'
    water_saturation_rock_type_chart_type = 'current'

    rock_type_colors = []

    def __init__(self):
        super().__init__()

        self.filename_fes_svod = None
        self.filename_sw = None
        self.water_saturation_rock_type_chart = Figure(figsize=(7, 7))
        self.rock_type_chart = Figure(figsize=(7, 7))
        self.FZI_chart = Figure(figsize=(7, 7))
        self.depth_column_name = tk.StringVar(value=self.FZI_work_book_columns[0])
        self.porv_column_name = tk.StringVar(value=self.FZI_work_book_columns[1])
        self.pron_column_name = tk.StringVar(value=self.FZI_work_book_columns[2])
        self.layer_column_name = tk.StringVar(value=self.FZI_work_book_columns[3])
        self.note_column_name = tk.StringVar(value=self.FZI_work_book_columns[4])

        '''menu start'''
        main_menu = Menu(self)

        help_menu = Menu(main_menu, tearoff=0)
        help_menu.add_command(label="Помощь", command=self.open_help_window)

        file_menu = Menu(main_menu, tearoff=0)
        file_menu.add_command(label="Выбрать Fes_svod...", command=self.choose_file_fes_svod)
        file_menu.add_command(label="Выбрать sw...", command=self.choose_file_sw)
        file_menu.add_command(label="Выбор столбцов", command=self.choose_column_from_work_book)

        self.config(menu=main_menu)
        main_menu.add_cascade(label='Файл', menu=file_menu)
        main_menu.add_cascade(label='Справка', menu=help_menu)
        '''menu end'''

        '''button frame start'''
        '''button fzi frame start'''
        frame_for_main_button = tk.Frame(self)

        fzi_frame = tk.Frame(master=frame_for_main_button)

        btn_calculate_rock_type = tk.Button(text="Рассчитать рок типы",
                                            command=self.calculate_rock_type,
                                            master=fzi_frame)

        interpretation = StringVar(fzi_frame)
        interpretation.set("Ручная настройка")  # default value
        interpretation_menu = OptionMenu(fzi_frame, interpretation, "Ручная настройка",
                                         "Алгоритм 1", "Алгоритм 2", "Алгоритм 3",
                                         command=self.choose_interpretation)

        type_rock_type_chart = StringVar(fzi_frame)
        type_rock_type_chart.set("RockType chart: log")  # default value
        type_rock_type_chart_menu = OptionMenu(fzi_frame, type_rock_type_chart, "RockType chart: log",
                                               "RockType chart: linear",
                                               command=self.choose_rock_type_chart)

        btn_water_saturation_rock_type = tk.Button(text="Рок тип/водонасыщенность",
                                                   command=self.calculate_water_saturation_rock_type,
                                                   master=frame_for_main_button)
        '''button fzi frame end'''

        '''button rock type frame start'''
        water_saturation_rock_type_chart_type = StringVar(frame_for_main_button)
        water_saturation_rock_type_chart_type.set("Остаточная")  # default value
        water_saturation_rock_type_chart_type_menu = OptionMenu(frame_for_main_button,
                                                                water_saturation_rock_type_chart_type,
                                                                "Остаточная", "Текущая",
                                                                command=self.choose_water_saturation_rock_type_chart_type)

        frame_for_main_button.pack()
        fzi_frame.pack(side=LEFT)
        interpretation_menu.pack(padx=10, pady=3, side=BOTTOM)
        type_rock_type_chart_menu.pack(padx=10, pady=3, side=BOTTOM)
        btn_calculate_rock_type.pack(padx=10, pady=3, side=BOTTOM)

        water_saturation_rock_type_chart_type_menu.pack(padx=10, pady=10, side=RIGHT)
        btn_water_saturation_rock_type.pack(padx=10, pady=10, side=RIGHT)
        '''button rock type frame end'''
        '''bottom frame end'''


    def choose_interpretation(self, interpretation):
        print(interpretation)

    def open_error_window(self, message):
        error_window = tk.Toplevel(app)
        label_error = tk.Label(text=message, master=error_window)
        label_error.pack()


    def open_help_window(self):
        help_window = tk.Toplevel(app)
        label_LKM = tk.Label(text='Добавить границу Рок Типа: для добавления гранциы рок типа '
                                  'нажмите левую кнопку мыши '
                                  'на графике', master=help_window)
        label_PKM = tk.Label(text='Удалить границу Рок Типа: для удаления последней добавленной гранциы рок типа '
                                  'нажмите правую кнопку мыши '
                                  'на графике', master=help_window)
        label_LKM.pack()
        label_PKM.pack()

    def choose_rock_type_chart(self, chart_type):
        if chart_type.__contains__('log'):
            self.rock_type_chart_scale = 'log'
        else:
            self.rock_type_chart_scale = 'linear'

    def choose_water_saturation_rock_type_chart_type(self, chart_type):
        if chart_type.__contains__('Остаточная'):
            self.water_saturation_rock_type_chart_type = 'residual'
        else:
            self.water_saturation_rock_type_chart_type = 'current'

    def choose_file_fes_svod(self):
        filename = fd.askopenfilename(title="Открыть файл", initialdir="/",
                                      filetypes=[("Excel files", ".xlsx .xls")])

        if filename:
            self.filename_fes_svod = filename
            self.rock_type_borders = []
            self.calculate_chart_auto_FZI(filename)
            self.draw_chart_auto_FZI()


    def choose_file_sw(self):
        filename = fd.askopenfilename(title="Открыть файл", initialdir="/",
                                      filetypes=[("Excel files", ".xlsx .xls")])

        if filename:
            self.filename_sw = filename

    def choose_column_from_work_book(self):
        column_choose_window = tk.Toplevel(self)

        edit_frame = tk.Frame(master=column_choose_window)
        label_frame = tk.Frame(master=edit_frame)
        label_depth_gis = tk.Label(text='Глубина отбора по гис(general)', master=label_frame)
        label_porv_gel = tk.Label(text='Пористость(PORV_gel)', master=label_frame)
        label_pron = tk.Label(text='Проницаемость(PRON_KL)', master=label_frame)
        label_layer = tk.Label(text='Пласты(general)', master=label_frame)
        label_note = tk.Label(text='Примечания(general)', master=label_frame)

        text_frame = tk.Frame(master=edit_frame)
        text_depth_gis = tk.Entry(master=text_frame, textvariable=self.depth_column_name)
        text_porv_gel = tk.Entry(master=text_frame, textvariable=self.porv_column_name)
        text_pron = tk.Entry(master=text_frame, textvariable=self.pron_column_name)
        text_layer = tk.Entry(master=text_frame, textvariable=self.layer_column_name)
        text_note = tk.Entry(master=text_frame, textvariable=self.note_column_name)

        edit_frame.pack()
        text_frame.pack(side=RIGHT)
        label_depth_gis.pack()
        label_porv_gel.pack()
        label_pron.pack()
        label_layer.pack()
        label_note.pack()

        label_frame.pack()
        text_depth_gis.pack()
        text_porv_gel.pack()
        text_pron.pack()
        text_layer.pack()
        text_note.pack()

        button_submit = tk.Button(text='Применить',
                                  master=column_choose_window,
                                  command=self.change_read_columns_work_book)
        button_submit.pack(side=BOTTOM)

    def change_read_columns_work_book(self):
        self.FZI_work_book_columns = [self.depth_column_name.get(),
                                      self.porv_column_name.get(),
                                      self.pron_column_name.get(),
                                      self.layer_column_name.get(),
                                      self.note_column_name.get()]

    def on_pick(self, event):
        if event.xdata == None:
            return
        if event.button == 3:
            if len(self.rock_type_borders) > 0:
                self.rock_type_borders.pop()
        else:
            self.rock_type_borders.append(event.xdata)
        self.draw_chart_auto_FZI()

    def calculate_chart_auto_FZI(self, filename):
        wb_fes = SheetReader(filename, self.FZI_work_book_columns[0],
                             self.FZI_work_book_columns[1],
                             self.FZI_work_book_columns[2],
                             self.FZI_work_book_columns[3],
                             self.FZI_work_book_columns[4], )
        wb_fes.run()
        self.FZI_chart_x = wb_fes.get_column_auto_FZI(['Log(FZI)'])
        self.FZI_chart_y = wb_fes.get_column_auto_FZI(['probability'])
        self.pron_por_fzi = wb_fes.get_column_auto_FZI(['Пористость', 'Проницаемость', 'Log(FZI)'])

    def calculate_rock_type(self):
        x_pts_modify = [-math.inf] + sorted(self.rock_type_borders) + [math.inf]
        self.dots_rock_type = []

        for x in range(0, len(x_pts_modify) - 1):
            rock_type_n = []

            for i in range(0, len(self.pron_por_fzi['Log(FZI)'])):
                if x_pts_modify[x] < self.pron_por_fzi['Log(FZI)'][i] < x_pts_modify[x + 1]:
                    rock_type_n.append([self.pron_por_fzi['Пористость'][i], self.pron_por_fzi['Проницаемость'][i]])

            self.dots_rock_type.append(rock_type_n)

        self.draw_chart_rock_type()

    def calculate_water_saturation_rock_type(self):

        if self.filename_sw is None:
            self.open_error_window('Выбирите файл sw.xlsx')
            return
        if self.filename_fes_svod is None:
            self.open_error_window('Выбирите файл FES_svod.xlsx')
            return

        """current_sw = SheetReader.get_column('C:/Users/kosac/PycharmProjects/RockType/Files/sw.xlsx', 'D')
        residual_sw = SheetReader.get_column('C:/FES_svod.xlsx', 'V')"""

        current_sw = SheetReader.get_column(self.filename_sw, 'D')
        residual_sw = SheetReader.get_column(self.filename_fes_svod, 'V')

        self.modify_current_sw = []
        for i in current_sw:
            if isinstance(i, float):
                self.modify_current_sw.append(i / 100)
        self.modify_current_sw = sorted(self.modify_current_sw)

        self.modify_residual_sw = []
        for i in residual_sw:
            if isinstance(i, float):
                self.modify_residual_sw.append(i / 100)
        self.modify_residual_sw = sorted(self.modify_residual_sw)
        self.draw_water_saturation_rock_type()

    def draw_chart_auto_FZI(self):
        self.delete_chart('fzi_chart')
        #оставить для очистки
        self.FZI_chart.clear()

        plot_fzi = self.FZI_chart.add_subplot(111)
        plot_fzi.plot(self.FZI_chart_x, self.FZI_chart_y)
        plot_fzi.set_ylabel('Probability')
        plot_fzi.set_xlabel('Log(FZI)')

        for x in sorted(self.rock_type_borders):
            plot_fzi.plot([x, x], [0, 1])

        canvas = FigureCanvasTkAgg(self.FZI_chart)
        canvas.mpl_connect('button_press_event', self.on_pick)
        canvas.get_tk_widget().pack(side=LEFT)
        canvas.draw()
        self.set_canvas_widget_name('fzi_chart')

        self.rock_type_colors = []
        for i in plot_fzi.get_lines():
            self.rock_type_colors.append(i.get_color())

    def draw_chart_rock_type(self):
        self.delete_chart('rock_type')
        #оставить для очистки
        self.rock_type_chart.clear()

        plot_rock_type = self.rock_type_chart.add_subplot(111)
        plot_rock_type.set_yscale(self.rock_type_chart_scale)
        plot_rock_type.set_ylabel('Проницаемость')
        plot_rock_type.set_xlabel('Пористость')

        for i in range(len(self.dots_rock_type)):
            if i < len(self.rock_type_borders):
                rock_type_value = str(round(self.rock_type_borders[i], 2))
            else:
                rock_type_value = str(math.inf)

            plot_rock_type.plot(0, 0, 'o', markersize=5, color=self.rock_type_colors[i],
                                label='Rock Type ' + str(i + 1) + ': ' + rock_type_value)
            plot_rock_type.legend(loc="lower right")

            for dots in self.dots_rock_type[i]:
                plot_rock_type.plot(dots[0], dots[1], 'o', markersize=4, color=self.rock_type_colors[i])

            d = pd.DataFrame(self.dots_rock_type[i], columns=['x1', 'y1'])
            x1 = numpy.array(d['x1'])
            y1 = numpy.array(d['y1'])

            z = numpy.polyfit(x1, y1, 1)
            p = numpy.poly1d(z)
            xx = []
            yy = []
            for i1 in range(len(x1)):
                if p(x1[i1]) > min(y1):
                    xx.append(x1[i1])
                    yy.append(p(x1[i1]))
            plot_rock_type.plot(sorted(xx), sorted(yy), "-", color=self.rock_type_colors[i],
                                label="y=%.6fx+(%.6f)" % (z[0], z[1]))

        canvas = FigureCanvasTkAgg(self.rock_type_chart)
        canvas.get_tk_widget().pack(side=RIGHT)
        canvas.draw()
        self.set_canvas_widget_name('rock_type')

    def draw_water_saturation_rock_type(self):
        self.delete_chart('water_saturation_rock_type')
        self.water_saturation_rock_type_chart.clear()

        if self.water_saturation_rock_type_chart_type == 'current':
            dots_x = self.modify_current_sw
        else:
            dots_x = self.modify_residual_sw

        plot_water_saturation_rock_type = self.water_saturation_rock_type_chart.add_subplot(111)
        for rock_type_number in range(1, len(self.rock_type_borders)+2):
            print(rock_type_number)
            if rock_type_number - 1 < len(self.rock_type_borders):
                rock_type_value = str(round(self.rock_type_borders[rock_type_number - 1], 2))
            else:
                rock_type_value = str(math.inf)

            plot_water_saturation_rock_type.plot(0, 1, 'o', markersize=0,
                                                 color=self.rock_type_colors[rock_type_number - 1],
                                                 label='Rock Type ' + str(rock_type_number) + ': ' + rock_type_value)
            plot_water_saturation_rock_type.legend(loc="lower right")

            for i in range(int(len(dots_x) / (len(self.rock_type_borders)+1) * rock_type_number),
                           int(len(dots_x) / (len(self.rock_type_borders)+1) * (rock_type_number + 1))):
                plot_water_saturation_rock_type.plot(i, rock_type_number, 'o',
                                                     markersize=1, color=self.rock_type_colors[rock_type_number - 1])

        canvas = FigureCanvasTkAgg(self.water_saturation_rock_type_chart)
        canvas.get_tk_widget().pack(side=RIGHT)
        canvas.draw()
        self.set_canvas_widget_name('water_saturation_rock_type')

    def set_canvas_widget_name(self, name):
        temp_elem = None
        for elem in self.pack_slaves():
            if elem.widgetName == 'canvas':
                temp_elem = elem

        if temp_elem != None:
            temp_elem.widgetName = name

    def delete_chart(self, name):
        for elem in self.pack_slaves():
            if elem.widgetName == name:
                elem.destroy()

    @staticmethod
    def trend(df):
        df = df.copy().sort_index()
        dates = df.index.to_julian_date().values[:, None]
        x = numpy.concatenate([numpy.ones_like(dates), dates], axis=1)
        y = df.values
        return pd.DataFrame(numpy.linalg.pinv(x.T.dot(x)).dot(x.T).dot(y).T,
                            df.columns, ['Constant', 'Trend'])

if __name__ == "__main__":
    app = App()
    app.mainloop()


